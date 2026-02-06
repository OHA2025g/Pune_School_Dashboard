"""Export routes for PDF and Excel"""
from fastapi import APIRouter, HTTPException, Depends, Response
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from typing import Optional, List
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart

from utils.auth import require_export_permission

router = APIRouter(prefix="/export", tags=["Export"])

# Database will be injected
db = None

def init_db(database):
    global db
    db = database

# Excel styling
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_excel_sheet(ws, headers, start_row=1):
    """Apply styling to Excel sheet"""
    # Set headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(str(headers[col-1]))
        for row in range(start_row + 1, ws.max_row + 1):
            try:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

def add_data_rows(ws, data, start_row=2):
    """Add data rows to Excel sheet"""
    for row_idx, row_data in enumerate(data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center' if isinstance(value, (int, float)) else 'left')

# ============== EXCEL EXPORTS ==============

@router.get("/excel/executive-summary")
async def export_executive_summary_excel(current_user: dict = Depends(require_export_permission)):
    """Export Executive Summary to Excel"""
    wb = Workbook()
    
    # Sheet 1: Overview KPIs
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    # Fetch data from all collections
    shi_data = await get_shi_data()
    
    # Add title
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "Maharashtra Education Dashboard - Executive Summary"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    ws1['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # KPI Summary
    headers = ["Domain", "Score", "Status", "Key Metric 1", "Key Metric 2", "Weight"]
    style_excel_sheet(ws1, headers, start_row=5)
    
    data = [
        ["Student Identity", shi_data.get("identity_index", 0), "Green" if shi_data.get("identity_index", 0) >= 85 else "Amber", 
         f"Aadhaar: {shi_data.get('aadhaar_pct', 0)}%", f"APAAR: {shi_data.get('apaar_pct', 0)}%", "25%"],
        ["Infrastructure", shi_data.get("infra_index", 0), "Green" if shi_data.get("infra_index", 0) >= 85 else "Amber",
         f"Classroom: {shi_data.get('classroom_health', 0)}%", f"Toilet: {shi_data.get('toilet_pct', 0)}%", "25%"],
        ["Teacher Quality", shi_data.get("teacher_index", 0), "Green" if shi_data.get("teacher_index", 0) >= 85 else "Amber",
         f"CTET: {shi_data.get('ctet_pct', 0)}%", f"NISHTHA: {shi_data.get('nishtha_pct', 0)}%", "25%"],
        ["Operational", shi_data.get("ops_index", 0), "Green" if shi_data.get("ops_index", 0) >= 85 else "Amber",
         f"Completion: {shi_data.get('completion_rate', 0)}%", f"Certification: {shi_data.get('cert_rate', 0)}%", "25%"],
        ["School Health Index", shi_data.get("shi", 0), shi_data.get("rag_status", "Red"), "", "", "100%"]
    ]
    add_data_rows(ws1, data, start_row=6)
    
    # Sheet 2: Block Rankings
    ws2 = wb.create_sheet("Block Rankings")
    block_data = await get_block_rankings()
    
    headers = ["Rank", "Block Name", "SHI Score", "Identity", "Infrastructure", "Teacher", "Operational", "RAG Status"]
    style_excel_sheet(ws2, headers)
    
    block_rows = [[b["rank"], b["block_name"], b["shi_score"], b["identity"], b["infra"], b["teacher"], b["ops"], b["rag"]] 
                  for b in block_data]
    add_data_rows(ws2, block_rows)
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=executive_summary_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

@router.get("/excel/{dashboard_name}")
async def export_dashboard_excel(dashboard_name: str, current_user: dict = Depends(require_export_permission)):
    """Export specific dashboard to Excel"""
    wb = Workbook()
    ws = wb.active
    
    # Map dashboard names to data fetchers
    dashboard_map = {
        "aadhaar": ("Aadhaar Analytics", fetch_aadhaar_data),
        "apaar": ("APAAR Status", fetch_apaar_data),
        "teacher": ("Teacher Analytics", fetch_teacher_data),
        "infrastructure": ("Infrastructure", fetch_infrastructure_data),
        "enrolment": ("Enrolment Analytics", fetch_enrolment_data),
        "classrooms-toilets": ("Classrooms & Toilets", fetch_classrooms_toilets_data),
        "dropbox": ("Dropbox Remarks", fetch_dropbox_data),
        "data-entry": ("Data Entry Status", fetch_data_entry_data),
        "age-enrolment": ("Age-wise Enrolment", fetch_age_enrolment_data),
        "ctteacher": ("CTTeacher Analytics", fetch_ctteacher_data),
    }
    
    if dashboard_name not in dashboard_map:
        raise HTTPException(status_code=404, detail=f"Dashboard '{dashboard_name}' not found")
    
    title, fetch_func = dashboard_map[dashboard_name]
    ws.title = title[:31]  # Excel sheet name limit
    
    # Add title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Maharashtra Education Dashboard - {title}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Fetch and add data
    data = await fetch_func()
    if data["headers"] and data["rows"]:
        style_excel_sheet(ws, data["headers"], start_row=5)
        add_data_rows(ws, data["rows"], start_row=6)
    
    # Add KPI summary sheet
    if data.get("kpis"):
        ws2 = wb.create_sheet("KPI Summary")
        kpi_headers = ["Metric", "Value", "Target", "Status"]
        style_excel_sheet(ws2, kpi_headers)
        add_data_rows(ws2, [[k["name"], k["value"], k.get("target", "-"), k.get("status", "-")] for k in data["kpis"]])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={dashboard_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ============== PDF EXPORTS ==============

@router.get("/pdf/executive-summary")
async def export_executive_summary_pdf(current_user: dict = Depends(require_export_permission)):
    """Export Executive Summary to PDF with charts"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, textColor=colors.grey)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=14, spaceBefore=15, spaceAfter=10)
    
    elements = []
    
    # Title
    elements.append(Paragraph("Maharashtra Education Dashboard", title_style))
    elements.append(Paragraph("Executive Summary Report", subtitle_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Fetch data
    shi_data = await get_shi_data()
    
    # School Health Index section
    elements.append(Paragraph("School Health Index (SHI)", heading_style))
    
    shi_table_data = [
        ["Metric", "Score", "Status"],
        ["School Health Index", f"{shi_data.get('shi', 0)}", shi_data.get('rag_status', 'Red')],
        ["Student Identity", f"{shi_data.get('identity_index', 0)}", "25% Weight"],
        ["Infrastructure", f"{shi_data.get('infra_index', 0)}", "25% Weight"],
        ["Teacher Quality", f"{shi_data.get('teacher_index', 0)}", "25% Weight"],
        ["Operational", f"{shi_data.get('ops_index', 0)}", "25% Weight"],
    ]
    
    shi_table = Table(shi_table_data, colWidths=[3*inch, 2*inch, 2*inch])
    shi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor("#E8F5E9") if shi_data.get('shi', 0) >= 85 else colors.HexColor("#FFEBEE")),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(shi_table)
    elements.append(Spacer(1, 20))
    
    # Key Statistics
    elements.append(Paragraph("Key Statistics", heading_style))
    stats_data = [
        ["Total Schools", "Total Students", "Total Teachers", "Total Classrooms", "Total Toilets"],
        [f"{shi_data.get('total_schools', 0):,}", f"{shi_data.get('total_students', 0):,}", 
         f"{shi_data.get('total_teachers', 0):,}", f"{shi_data.get('total_classrooms', 0):,}",
         f"{shi_data.get('total_toilets', 0):,}"]
    ]
    stats_table = Table(stats_data, colWidths=[1.5*inch]*5)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(stats_table)
    elements.append(PageBreak())
    
    # Block Rankings
    elements.append(Paragraph("Block-wise Performance Rankings", heading_style))
    block_data = await get_block_rankings()
    
    block_table_data = [["Rank", "Block", "SHI", "Identity", "Infra", "Teacher", "Ops", "Status"]]
    for b in block_data[:15]:
        block_table_data.append([
            str(b["rank"]), b["block_name"], f"{b['shi_score']}", 
            f"{b['identity']}", f"{b['infra']}", f"{b['teacher']}", f"{b['ops']}", b["rag"]
        ])
    
    block_table = Table(block_table_data, colWidths=[0.5*inch, 1.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
    block_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(block_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=executive_summary_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

@router.get("/pdf/{dashboard_name}")
async def export_dashboard_pdf(dashboard_name: str, current_user: dict = Depends(require_export_permission)):
    """Export specific dashboard to PDF"""
    dashboard_map = {
        "aadhaar": ("Aadhaar Analytics", fetch_aadhaar_data),
        "apaar": ("APAAR Status", fetch_apaar_data),
        "teacher": ("Teacher Analytics", fetch_teacher_data),
        "infrastructure": ("Infrastructure", fetch_infrastructure_data),
        "enrolment": ("Enrolment Analytics", fetch_enrolment_data),
        "classrooms-toilets": ("Classrooms & Toilets", fetch_classrooms_toilets_data),
        "dropbox": ("Dropbox Remarks", fetch_dropbox_data),
        "data-entry": ("Data Entry Status", fetch_data_entry_data),
        "age-enrolment": ("Age-wise Enrolment", fetch_age_enrolment_data),
        "ctteacher": ("CTTeacher Analytics", fetch_ctteacher_data),
    }
    
    if dashboard_name not in dashboard_map:
        raise HTTPException(status_code=404, detail=f"Dashboard '{dashboard_name}' not found")
    
    title, fetch_func = dashboard_map[dashboard_name]
    data = await fetch_func()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=15)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=1, textColor=colors.grey)
    heading_style = ParagraphStyle('Heading', parent=styles['Heading2'], fontSize=12, spaceBefore=10, spaceAfter=8)
    
    elements = []
    
    # Title
    elements.append(Paragraph(f"Maharashtra Education Dashboard - {title}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 15))
    
    # KPIs
    if data.get("kpis"):
        elements.append(Paragraph("Key Performance Indicators", heading_style))
        kpi_data = [["Metric", "Value", "Status"]]
        for kpi in data["kpis"][:10]:
            kpi_data.append([kpi["name"], str(kpi["value"]), kpi.get("status", "-")])
        
        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch, 1.5*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 15))
    
    # Data table
    if data.get("headers") and data.get("rows"):
        elements.append(Paragraph("Detailed Data", heading_style))
        table_data = [data["headers"]] + data["rows"][:30]  # Limit rows for PDF
        
        col_width = min(1.2*inch, 10*inch / len(data["headers"]))
        detail_table = Table(table_data, colWidths=[col_width] * len(data["headers"]))
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        elements.append(detail_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={dashboard_name}_{datetime.now().strftime('%Y%m%d')}.pdf"}
    )

# ============== DATA FETCHERS ==============

async def get_shi_data():
    """Fetch SHI and overview data"""
    # Get infrastructure data
    ct_pipeline = [{"$group": {"_id": None, "schools": {"$sum": 1}, "classrooms": {"$sum": "$classrooms_instructional"},
                               "good": {"$sum": {"$add": ["$pucca_good", "$part_pucca_good"]}},
                               "toilets": {"$sum": {"$add": ["$boys_toilets_total", "$girls_toilets_total"]}},
                               "func_toilets": {"$sum": {"$add": ["$boys_toilets_functional", "$girls_toilets_functional"]}}}}]
    ct_result = await db.classrooms_toilets.aggregate(ct_pipeline).to_list(1)
    ct = ct_result[0] if ct_result else {}
    
    # Get APAAR data
    apaar_pipeline = [{"$group": {"_id": None, "students": {"$sum": "$total_student"}, "generated": {"$sum": "$total_generated"}}}]
    apaar_result = await db.apaar_status.aggregate(apaar_pipeline).to_list(1)
    apaar = apaar_result[0] if apaar_result else {}
    
    # Get teacher data
    teacher_count = await db.ctteacher.count_documents({})
    
    # Calculate indices
    classroom_health = round(ct.get("good", 0) / max(ct.get("classrooms", 1), 1) * 100, 1)
    toilet_pct = round(ct.get("func_toilets", 0) / max(ct.get("toilets", 1), 1) * 100, 1)
    apaar_pct = round(apaar.get("generated", 0) / max(apaar.get("students", 1), 1) * 100, 1)
    
    infra_index = round((classroom_health * 0.5 + toilet_pct * 0.5), 1)
    identity_index = round(apaar_pct * 0.6 + 40, 1)  # Simplified
    teacher_index = 30.0  # Placeholder
    ops_index = 50.0  # Placeholder
    
    shi = round((identity_index + infra_index + teacher_index + ops_index) / 4, 1)
    
    return {
        "shi": shi,
        "rag_status": "Green" if shi >= 85 else "Amber" if shi >= 70 else "Red",
        "identity_index": identity_index,
        "infra_index": infra_index,
        "teacher_index": teacher_index,
        "ops_index": ops_index,
        "aadhaar_pct": 85,
        "apaar_pct": apaar_pct,
        "classroom_health": classroom_health,
        "toilet_pct": toilet_pct,
        "ctet_pct": 6.6,
        "nishtha_pct": 32.3,
        "completion_rate": 99.9,
        "cert_rate": 45,
        "total_schools": ct.get("schools", 0),
        "total_students": apaar.get("students", 0),
        "total_teachers": teacher_count,
        "total_classrooms": ct.get("classrooms", 0),
        "total_toilets": ct.get("toilets", 0)
    }

async def get_block_rankings():
    """Fetch block-wise rankings"""
    pipeline = [
        {"$group": {
            "_id": "$block_name",
            "schools": {"$sum": 1},
            "classrooms": {"$sum": "$classrooms_instructional"},
            "good": {"$sum": {"$add": ["$pucca_good", "$part_pucca_good"]}}
        }},
        {"$sort": {"schools": -1}}
    ]
    blocks = await db.classrooms_toilets.aggregate(pipeline).to_list(30)
    
    result = []
    for i, b in enumerate(blocks):
        health = round(b["good"] / max(b["classrooms"], 1) * 100, 1)
        shi_score = round((health * 0.3 + 85 * 0.3 + 30 * 0.2 + 50 * 0.2), 1)
        result.append({
            "rank": i + 1,
            "block_name": b["_id"],
            "shi_score": shi_score,
            "identity": 85,
            "infra": health,
            "teacher": 30,
            "ops": 50,
            "rag": "Green" if shi_score >= 85 else "Amber" if shi_score >= 70 else "Red"
        })
    
    result.sort(key=lambda x: x["shi_score"], reverse=True)
    for i, r in enumerate(result):
        r["rank"] = i + 1
    
    return result

async def fetch_aadhaar_data():
    """Fetch Aadhaar analytics data"""
    data = await db.aadhaar_analytics.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Total Students", "Aadhaar Available", "Coverage %", "Name Mismatch", "MBU Pending"]
    rows = []
    for d in data[:100]:
        rows.append([
            d.get("block_name", ""),
            d.get("total_students", 0),
            d.get("aadhaar_available", 0),
            round(d.get("aadhaar_available", 0) / max(d.get("total_students", 1), 1) * 100, 1),
            d.get("name_match_failed", 0),
            d.get("mbu_pending", 0)
        ])
    
    total = sum(d.get("total_students", 0) for d in data)
    aadhaar = sum(d.get("aadhaar_available", 0) for d in data)
    kpis = [
        {"name": "Total Students", "value": total, "status": "Info"},
        {"name": "Aadhaar Coverage", "value": f"{round(aadhaar/max(total,1)*100, 1)}%", "status": "Good" if aadhaar/max(total,1) > 0.9 else "Warning"}
    ]
    
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_apaar_data():
    """Fetch APAAR status data"""
    data = await db.apaar_status.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Total Students", "Generated", "Generation %", "Pending", "Not Applied"]
    rows = []
    for d in data[:100]:
        rows.append([
            d.get("block_name", ""),
            d.get("total_student", 0),
            d.get("total_generated", 0),
            round(d.get("total_generated", 0) / max(d.get("total_student", 1), 1) * 100, 1),
            d.get("total_pending", 0),
            d.get("total_not_applied", 0)
        ])
    
    total = sum(d.get("total_student", 0) for d in data)
    generated = sum(d.get("total_generated", 0) for d in data)
    kpis = [
        {"name": "Total Students", "value": total, "status": "Info"},
        {"name": "APAAR Generated", "value": generated, "status": "Info"},
        {"name": "Generation Rate", "value": f"{round(generated/max(total,1)*100, 1)}%", "status": "Good" if generated/max(total,1) > 0.85 else "Warning"}
    ]
    
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_teacher_data():
    """Fetch teacher analytics data"""
    data = await db.teacher_analytics.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Teachers CY", "Teachers PY", "Growth", "CTET", "CWSN Trained"]
    rows = [[d.get("block_name", ""), d.get("teachers_cy", 0), d.get("teachers_py", 0),
             d.get("teachers_cy", 0) - d.get("teachers_py", 0), d.get("ctet_cy", 0), d.get("cwsn_trained", 0)]
            for d in data[:100]]
    
    kpis = [{"name": "Total Teachers CY", "value": sum(d.get("teachers_cy", 0) for d in data), "status": "Info"}]
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_infrastructure_data():
    """Fetch infrastructure data"""
    data = await db.infrastructure_analytics.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Schools", "Tap Water", "Water Purifier", "Water Tested", "Ramp"]
    rows = [[d.get("block_name", ""), 1, 1 if d.get("tap_water") else 0, 1 if d.get("water_purifier") else 0,
             1 if d.get("water_tested") else 0, 1 if d.get("ramp") else 0] for d in data[:100]]
    
    kpis = [{"name": "Total Schools", "value": len(data), "status": "Info"}]
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_enrolment_data():
    """Fetch enrolment data"""
    data = await db.enrolment_analytics.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Total Enrolment", "Boys", "Girls", "Girls %"]
    rows = [[d.get("block_name", ""), d.get("total_enrolment", 0), d.get("boys", 0), d.get("girls", 0),
             round(d.get("girls", 0) / max(d.get("total_enrolment", 1), 1) * 100, 1)] for d in data[:100]]
    
    total = sum(d.get("total_enrolment", 0) for d in data)
    girls = sum(d.get("girls", 0) for d in data)
    kpis = [
        {"name": "Total Enrolment", "value": total, "status": "Info"},
        {"name": "Girls %", "value": f"{round(girls/max(total,1)*100, 1)}%", "status": "Good" if girls/max(total,1) > 0.48 else "Warning"}
    ]
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_classrooms_toilets_data():
    """Fetch classrooms and toilets data"""
    data = await db.classrooms_toilets.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["School", "Block", "Classrooms", "Good Condition", "Toilets", "Functional"]
    rows = [[d.get("school_name", "")[:30], d.get("block_name", ""), d.get("classrooms_instructional", 0),
             d.get("pucca_good", 0) + d.get("part_pucca_good", 0),
             d.get("boys_toilets_total", 0) + d.get("girls_toilets_total", 0),
             d.get("boys_toilets_functional", 0) + d.get("girls_toilets_functional", 0)] for d in data[:100]]
    
    total_classrooms = sum(d.get("classrooms_instructional", 0) for d in data)
    good = sum(d.get("pucca_good", 0) + d.get("part_pucca_good", 0) for d in data)
    kpis = [
        {"name": "Total Schools", "value": len(data), "status": "Info"},
        {"name": "Classroom Health", "value": f"{round(good/max(total_classrooms,1)*100, 1)}%", "status": "Good"}
    ]
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_dropbox_data():
    """Fetch dropbox remarks data"""
    data = await db.dropbox_analytics.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Total Remarks", "Dropout", "Migration", "Class 12 Passed", "Wrong Entry"]
    rows = [[d.get("block_name", ""), d.get("total_remarks", 0), d.get("dropout", 0),
             d.get("migration", 0), d.get("class12_passed", 0), d.get("wrong_entry", 0)] for d in data[:100]]
    
    kpis = [{"name": "Total Remarks", "value": sum(d.get("total_remarks", 0) for d in data), "status": "Info"}]
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_data_entry_data():
    """Fetch data entry status"""
    data = await db.data_entry_status.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Total Students", "Completed", "Pending", "Certified"]
    rows = [[d.get("block_name", ""), d.get("total_students", 0), d.get("completed", 0),
             d.get("pending", 0), "Yes" if d.get("certified") else "No"] for d in data[:100]]
    
    kpis = [{"name": "Total Students", "value": sum(d.get("total_students", 0) for d in data), "status": "Info"}]
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_age_enrolment_data():
    """Fetch age-wise enrolment data"""
    data = await db.age_enrolment.find({}, {"_id": 0}).to_list(1000)
    if not data:
        return {"headers": [], "rows": [], "kpis": []}
    
    headers = ["Block", "Age Group", "Boys", "Girls", "Total"]
    rows = [[d.get("block_name", ""), d.get("age_group", ""), d.get("boys", 0),
             d.get("girls", 0), d.get("boys", 0) + d.get("girls", 0)] for d in data[:100]]
    
    kpis = [{"name": "Total Records", "value": len(data), "status": "Info"}]
    return {"headers": headers, "rows": rows, "kpis": kpis}

async def fetch_ctteacher_data():
    """Fetch CTTeacher data"""
    pipeline = [
        {"$group": {
            "_id": "$block_name",
            "count": {"$sum": 1},
            "ctet": {"$sum": {"$cond": [{"$eq": ["$ctet_qualified", True]}, 1, 0]}},
            "nishtha": {"$sum": {"$cond": [{"$eq": ["$nishtha_completed", True]}, 1, 0]}}
        }},
        {"$sort": {"count": -1}}
    ]
    data = await db.ctteacher.aggregate(pipeline).to_list(100)
    
    headers = ["Block", "Teachers", "CTET Qualified", "CTET %", "NISHTHA Completed", "NISHTHA %"]
    rows = [[d["_id"], d["count"], d["ctet"], round(d["ctet"]/max(d["count"],1)*100, 1),
             d["nishtha"], round(d["nishtha"]/max(d["count"],1)*100, 1)] for d in data]
    
    total = sum(d["count"] for d in data)
    ctet = sum(d["ctet"] for d in data)
    kpis = [
        {"name": "Total Teachers", "value": total, "status": "Info"},
        {"name": "CTET Rate", "value": f"{round(ctet/max(total,1)*100, 1)}%", "status": "Warning" if ctet/max(total,1) < 0.5 else "Good"}
    ]
    return {"headers": headers, "rows": rows, "kpis": kpis}
